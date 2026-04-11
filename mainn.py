# %% [Cell 1] - Imports
import pandas as pd
import numpy as np
import yfinance as yf

from sklearn.pipeline import Pipeline
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.tree import DecisionTreeClassifier
from sklearn.linear_model import LogisticRegression

from newsapi import NewsApiClient
from textblob import TextBlob

# %% [Cell 2] - Download Stock Data
stocks = ["TCS.NS", "INFY.NS", "RELIANCE.NS", "HDFCBANK.NS"]

df_list = []

for stock in stocks:
    temp_df = yf.download(stock, start="2022-01-01", end="2025-01-01")

    temp_df.reset_index(inplace=True)

    # FIX: Flatten MultiIndex columns if present (yfinance returns MultiIndex)
    if isinstance(temp_df.columns, pd.MultiIndex):
        temp_df.columns = [col[0] for col in temp_df.columns]

    # Rename to standard column names
    # After reset_index, columns are: Date, Close, High, Low, Open, Volume (order may vary)
    # Keep whatever yfinance gives us, just ensure we have the right ones
    temp_df['Stock'] = stock

    df_list.append(temp_df)

# Combine all stocks
df = pd.concat(df_list, ignore_index=True)

print("✓ Data downloaded successfully")
print(df.head())

# %% [Cell 3] - Data Info
print(f"\nShape: {df.shape}")
df.info()

# %% [Cell 4] - Data Cleaning
# Flatten columns if still MultiIndex (safety check)
if isinstance(df.columns, pd.MultiIndex):
    df.columns = df.columns.get_level_values(0)

# Remove duplicate columns
df = df.loc[:, ~df.columns.duplicated()]

# Check what columns we have
print("Available columns:", df.columns.tolist())

# Keep required columns (only select if they exist) — IMPORTANT: include 'Stock'!
required_cols = ['Date', 'Open', 'High', 'Low', 'Close', 'Volume', 'Stock']
available_cols = [col for col in required_cols if col in df.columns]
df = df[available_cols]

# Convert data types
df['Date'] = pd.to_datetime(df['Date'])
df[['Open', 'High', 'Low', 'Close', 'Volume']] = df[
    ['Open', 'High', 'Low', 'Close', 'Volume']
].astype('float32')

print("\n✓ Columns after cleaning:", df.columns.tolist())
df.head()

# %% [Cell 5] - Sort and Drop Nulls
# Sort by Date
df.sort_values('Date', inplace=True)
df.reset_index(drop=True, inplace=True)

# Drop null values
df.dropna(inplace=True)

print(f"✓ Shape after sorting & cleaning: {df.shape}")
df.head()

# %% [Cell 6] - Feature Engineering
# Moving averages
df['MA_5'] = df['Close'].rolling(5).mean()
df['MA_10'] = df['Close'].rolling(10).mean()

# Lag features
df['Lag_1'] = df['Close'].shift(1)
df['Lag_2'] = df['Close'].shift(2)

# Volatility
df['Volatility'] = df['Close'].rolling(5).std()

# Return
df['Return'] = df['Close'].pct_change()

# Volume MA
df['Volume_MA'] = df['Volume'].rolling(5).mean()

# Target (next day up/down)
df['Target'] = (df['Close'].shift(-1) > df['Close']).astype(int)

# Additional features
df['MA_diff'] = df['MA_5'] - df['MA_10']
df['Price_Change'] = df['Close'] - df['Lag_1']
df['Volume_Change'] = df['Volume'].pct_change()

# Remove NaN rows (MUST be LAST after all feature engineering)
df.dropna(inplace=True)

print(f"✓ Shape after feature engineering: {df.shape}")
print(f"✓ Columns: {df.columns.tolist()}")
print(f"✓ 'Stock' column exists: {'Stock' in df.columns}")
print(df.head())

# %% [Cell 7] - Sentiment Analysis
newsapi = NewsApiClient(api_key='4cb5f1c4b1ba44829d55c3bb4207590c')

def get_sentiment(stock_name):
    try:
        articles = newsapi.get_everything(q=stock_name, language='en', page_size=5)

        scores = []
        for article in articles['articles']:
            title = article['title']
            score = TextBlob(title).sentiment.polarity
            scores.append(score)

        if len(scores) == 0:
            return 0

        return sum(scores) / len(scores)
    except Exception as e:
        print(f"  ⚠️ NewsAPI error for {stock_name}: {e}")
        return 0

# Get sentiment scores for each stock
sentiment_scores = {}

for stock in stocks:
    sentiment = get_sentiment(stock)
    sentiment_scores[stock] = sentiment
    print(f"{stock}: {sentiment:.4f}")

# Create sentiment dataframe
sentiment_df = pd.DataFrame(list(sentiment_scores.items()), columns=['Stock', 'Sentiment'])
print("\nSentiment Scores:")
print(sentiment_df)

# Add sentiment column to dataframe (using average sentiment score)
avg_sentiment = sentiment_df['Sentiment'].mean()
df['Sentiment'] = avg_sentiment

# %% [Cell 8] - Train/Test Split + Model Training
features = [
    'MA_5', 'MA_10',
    'Lag_1', 'Lag_2',
    'Volatility',
    'Return',
    'Volume_MA',
    'MA_diff',
    'Price_Change',
    'Volume_Change'
]

X = df[features]
y = df['Target']

# Split data
X_train, X_test, y_train, y_test = train_test_split(
    X, y,
    test_size=0.2,
    random_state=42
)

print(f"✓ Data prepared: X_train shape {X_train.shape}, X_test shape {X_test.shape}")

# BUILD AND TRAIN PIPELINE
pipeline = Pipeline([
    ('scaler', StandardScaler()),
    ('classifier', RandomForestClassifier(
        n_estimators=200,
        max_depth=15,
        min_samples_split=5,
        min_samples_leaf=2,
        max_features='sqrt',
        random_state=42,
        n_jobs=-1
    ))
])

# Train the pipeline
try:
    print("Training Random Forest pipeline...")
    pipeline.fit(X_train, y_train)

    train_score = pipeline.score(X_train, y_train)
    test_score = pipeline.score(X_test, y_test)

    print(f"✓ Pipeline trained successfully!")
    print(f"✓ Training Accuracy: {train_score:.4f}")
    print(f"✓ Test Accuracy: {test_score:.4f}")

except Exception as e:
    print(f"✗ Error training pipeline: {str(e)}")

# %% [Cell 9] - Predictions & Results
pred = pipeline.predict(X_test)
conf = pipeline.predict_proba(X_test)[:, 1]

# Create results with proper STOCK + DATE mapping (using test indices)
results = pd.DataFrame({
    'Actual': y_test.values,
    'Prediction': pred,
    'Confidence': conf
})

results['Stock'] = df.loc[y_test.index, 'Stock'].values
results['Date'] = df.loc[y_test.index, 'Date'].values

# Signal generation (adjusted thresholds)
results['Signal'] = results.apply(
    lambda row: "BUY" if row['Confidence'] > 0.65
    else ("SELL" if row['Confidence'] < 0.35 else "HOLD"),
    axis=1
)

results['Correct'] = (results['Prediction'] == results['Actual']).astype(int)

# Print results summary
print("=" * 60)
print("PREDICTION RESULTS")
print("=" * 60)
print(f"✓ Model Accuracy: {results['Correct'].mean():.4f}")
print(f"✓ Total predictions: {len(results)}")
print(f"\n✓ BUY Signals:  {(results['Signal'] == 'BUY').sum()}")
print(f"✓ SELL Signals: {(results['Signal'] == 'SELL').sum()}")
print(f"✓ HOLD Signals: {(results['Signal'] == 'HOLD').sum()}")
print(f"\n✓ Confidence — Min: {results['Confidence'].min():.4f}, "
      f"Max: {results['Confidence'].max():.4f}, "
      f"Mean: {results['Confidence'].mean():.4f}")
print("\n✓ Sample predictions:")
print(results[['Stock', 'Date', 'Prediction', 'Confidence', 'Signal']].head(15))

results.head()

# %% [Cell 10] - Export to Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

output_file = "final.xlsx"
results_export = results[['Date', 'Stock', 'Prediction', 'Signal', 'Confidence', 'Actual', 'Correct']].copy()

# Format Date — remove time, keep only date
results_export['Date'] = pd.to_datetime(results_export['Date']).dt.strftime('%Y-%m-%d')

# Export
results_export.to_excel(output_file, index=False)

# Format workbook
wb = openpyxl.load_workbook(output_file)
ws = wb.active

ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 10

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal="center")

wb.save(output_file)

print(f"✓ Results exported to {output_file}")
print(f"✓ Total rows exported: {len(results_export)}")

# %% [Cell 11] - Model Comparison
# Decision Tree
tree_model = DecisionTreeClassifier()
tree_model.fit(X_train, y_train)
tree_pred = tree_model.predict(X_test)
tree_acc = (tree_pred == y_test).mean()
print(f"Decision Tree Accuracy: {tree_acc:.4f}")

# Random Forest (already trained)
rf_acc = results['Correct'].mean()
print(f"Random Forest Accuracy: {rf_acc:.4f}")

# Logistic Regression
log_model = LogisticRegression(max_iter=500)
log_model.fit(X_train, y_train)
log_pred = log_model.predict(X_test)
log_acc = (log_pred == y_test).mean()
print(f"Logistic Regression Accuracy: {log_acc:.4f}")

# Comparison table
comparison = pd.DataFrame({
    'Model': ['Decision Tree', 'Random Forest', 'Logistic Regression'],
    'Accuracy': [tree_acc, rf_acc, log_acc]
})

print("\n" + "=" * 40)
print("MODEL COMPARISON")
print("=" * 40)
print(comparison)

comparison.to_excel("model_comparison.xlsx", index=False)
print("\n✓ Comparison exported to model_comparison.xlsx")
